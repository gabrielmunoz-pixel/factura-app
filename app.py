import streamlit as st
import base64
import json
import io
import re
import time

import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="FacturaAI", page_icon="🧾", layout="wide")

st.markdown("""
<style>
.block-container { padding-top: 2rem; }
.stButton > button {
    background: #0a0a0f !important; color: white !important;
    border-radius: 8px !important; font-weight: 700 !important; width: 100%;
}
.stButton > button:hover { background: #e8521a !important; }
.info-card { background:#f5f2eb; border-radius:8px; padding:12px 14px; margin-bottom:8px; }
.info-label { font-size:10px; text-transform:uppercase; color:#8a8680; margin-bottom:2px; }
.info-value { font-size:14px; font-weight:700; }
.total-grand { font-size:20px; font-weight:800; color:#e8521a; border-top:2px solid #0a0a0f; padding-top:10px; margin-top:6px; display:flex; justify-content:space-between; }
.lote-card { background:#fff; border:1.5px solid #ddd9d0; border-radius:10px; padding:14px 16px; margin-bottom:10px; }
.lote-ok { border-left: 4px solid #2a6b4f; }
.lote-err { border-left: 4px solid #e8521a; }
</style>
""", unsafe_allow_html=True)

# ── Prompt ────────────────────────────────────────────────────────────────────
PROMPT = """Eres un experto en lectura de facturas. Analiza esta factura con maxima precision.
Responde SOLO con JSON puro, sin texto adicional, sin backticks.

{
  "emisor": {"nombre":"","rut":"","direccion":"","giro":""},
  "receptor": {"nombre":"","rut":"","direccion":"","comuna":"","ciudad":"","giro":""},
  "factura": {"numero":"","fecha":"","tipo":"","condicion_venta":"","vendedor":""},
  "items": [{"codigo":"","descripcion":"","unidad":"","cantidad":0,"precio_unitario":0,"descuento":0,"total":0}],
  "totales": {"neto":0,"exento":0,"iva":0,"otros_impuestos":[],"total":0},
  "notas":"",
  "despacho":""
}

Lee cada linea con cuidado. Montos son numeros sin puntos ni simbolo pesos. Usa comillas vacias o 0 si no encuentras el dato."""

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_json(text):
    clean = re.sub(r"```json|```", "", text).strip()
    try:
        return json.loads(clean)
    except Exception:
        match = re.search(r"\{[\s\S]*\}", clean)
        if match:
            return json.loads(match.group())
        raise ValueError("No se encontro JSON en la respuesta")

def fmt_money(n):
    try:
        return "$" + "{:,}".format(int(n)).replace(",", ".")
    except Exception:
        return "-"

def call_claude(api_key, file_bytes, media_type):
    client = anthropic.Anthropic(api_key=api_key)
    b64 = base64.standard_b64encode(file_bytes).decode()
    if media_type == "application/pdf":
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
            {"type": "text", "text": PROMPT}
        ]
    else:
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
            {"type": "text", "text": PROMPT}
        ]
    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4096,
        messages=[{"role": "user", "content": content}]
    )
    return response.content[0].text

def factura_to_rows(data, filename=""):
    """Convierte una factura en filas denormalizadas (una por item)."""
    emisor   = data.get("emisor", {})
    receptor = data.get("receptor", {})
    factura  = data.get("factura", {})
    totales  = data.get("totales", {})
    items    = data.get("items", [])

    base = {
        "archivo":            filename,
        "folio":              factura.get("numero", ""),
        "fecha":              factura.get("fecha", ""),
        "tipo_documento":     factura.get("tipo", ""),
        "condicion_venta":    factura.get("condicion_venta", ""),
        "vendedor":           factura.get("vendedor", ""),
        "emisor_nombre":      emisor.get("nombre", ""),
        "emisor_rut":         emisor.get("rut", ""),
        "emisor_giro":        emisor.get("giro", ""),
        "receptor_nombre":    receptor.get("nombre", ""),
        "receptor_rut":       receptor.get("rut", ""),
        "receptor_comuna":    receptor.get("comuna", ""),
        "receptor_ciudad":    receptor.get("ciudad", ""),
        "receptor_giro":      receptor.get("giro", ""),
        "monto_neto":         totales.get("neto", 0),
        "monto_iva":          totales.get("iva", 0),
        "monto_total":        totales.get("total", 0),
        "notas":              data.get("notas", ""),
    }

    if not items:
        return [base]

    rows = []
    for item in items:
        row = dict(base)
        row["item_codigo"]         = item.get("codigo", "")
        row["item_descripcion"]    = item.get("descripcion", "")
        row["item_unidad"]         = item.get("unidad", "")
        row["item_cantidad"]       = item.get("cantidad", 0)
        row["item_precio_unitario"]= item.get("precio_unitario", 0)
        row["item_descuento"]      = item.get("descuento", 0)
        row["item_total"]          = item.get("total", 0)
        rows.append(row)
    return rows

def build_excel_single(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Factura"
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    alt_fill = PatternFill("solid", start_color="DEEAF1")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    bold = Font(bold=True, name="Calibri", size=10)
    normal = Font(name="Calibri", size=10)

    ws.merge_cells("A1:G1")
    ws["A1"] = "FACTURA {} - {}".format(
        data.get("factura", {}).get("numero", ""),
        data.get("emisor", {}).get("nombre", "")
    )
    ws["A1"].font = Font(bold=True, name="Calibri", size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    meta = [
        ("Fecha",        data.get("factura", {}).get("fecha", "")),
        ("Receptor",     data.get("receptor", {}).get("nombre", "")),
        ("RUT Receptor", data.get("receptor", {}).get("rut", "")),
        ("Vendedor",     data.get("factura", {}).get("vendedor", "")),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=2, value=v).font = normal

    headers = ["Codigo", "Descripcion", "U/M", "Cantidad", "Precio Unit.", "Dto.", "Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center"); c.border = bdr

    for i, item in enumerate(data.get("items", []), 9):
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        vals = [item.get("codigo",""), item.get("descripcion",""), item.get("unidad",""),
                item.get("cantidad",0), item.get("precio_unitario",0), item.get("descuento",0), item.get("total",0)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = normal; c.fill = fill; c.border = bdr
            if col in (4, 5, 6, 7):
                c.alignment = Alignment(horizontal="right")

    t = data.get("totales", {})
    trow = len(data.get("items", [])) + 10
    for j, (label, val) in enumerate([("Neto", t.get("neto",0)), ("Exento", t.get("exento",0)),
                                       ("IVA 19%", t.get("iva",0)), ("TOTAL", t.get("total",0))], trow):
        if not val: continue
        is_grand = label == "TOTAL"
        color = "C44A1A" if is_grand else "000000"
        c_l = ws.cell(row=j, column=5, value=label)
        c_v = ws.cell(row=j, column=7, value=val)
        c_l.font = Font(bold=True, name="Calibri", color=color)
        c_v.font = Font(bold=True, name="Calibri", color=color)
        c_l.alignment = Alignment(horizontal="right")
        c_v.alignment = Alignment(horizontal="right")
        c_v.border = bdr

    if data.get("notas"):
        ws.cell(row=trow+6, column=1, value="NOTA: {}".format(data["notas"])).font = Font(italic=True, color="C44A1A")

    for col, w in zip(range(1, 8), [14, 42, 7, 10, 14, 8, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()

def build_excel_lote(df):
    """Excel base de datos para lote de facturas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Datos"
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    alt_fill = PatternFill("solid", start_color="DEEAF1")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    normal = Font(name="Calibri", size=10)

    cols = list(df.columns)
    for col_idx, col_name in enumerate(cols, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center"); c.border = bdr
    ws.row_dimensions[1].height = 20

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = normal; c.fill = fill; c.border = bdr
            if col_idx > len(cols) - 7:
                c.alignment = Alignment(horizontal="right")

    # Auto width
    for col_idx, col_name in enumerate(cols, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) > 0 else 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()

def build_csv_single(data):
    lines = ["Codigo,Descripcion,Unidad,Cantidad,Precio Unitario,Total"]
    for item in data.get("items", []):
        lines.append('"{}","{}","{}",{},{},{}'.format(
            item.get("codigo",""), item.get("descripcion",""), item.get("unidad",""),
            item.get("cantidad",0), item.get("precio_unitario",0), item.get("total",0)
        ))
    return "\n".join(lines)

# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown("<h1 style='font-size:28px;font-weight:800'><span style='color:#e8521a'>Factura</span>AI</h1>", unsafe_allow_html=True)
st.caption("Lector inteligente de facturas · Powered by Claude")

api_key = st.secrets.get("ANTHROPIC_API_KEY", "")

with st.sidebar:
    st.markdown("### FacturaAI")
    st.caption("Sube una foto o PDF de tu factura y presiona Analizar.")
    st.markdown("**Formatos soportados:** JPG, PNG, WEBP, PDF")
    if not api_key:
        st.error("API Key no configurada en Secrets.")

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📄  Factura Individual", "📦  Lote de Facturas"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: Factura individual
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    col1, col2 = st.columns([1, 1.6], gap="large")

    with col1:
        st.markdown("#### Subir Factura")
        uploaded = st.file_uploader("Selecciona imagen o PDF", type=["jpg","jpeg","png","webp","pdf"], key="single")
        if uploaded:
            if uploaded.type.startswith("image/"):
                st.image(uploaded, use_container_width=True)
            else:
                st.success("PDF cargado: **{}**".format(uploaded.name))

        analyze = st.button("Analizar Factura", disabled=not uploaded or not api_key, key="btn_single")

    with col2:
        st.markdown("#### Datos Extraidos")

        if analyze and uploaded and api_key:
            with st.spinner("Claude esta leyendo tu factura..."):
                try:
                    file_bytes = uploaded.read()
                    media_type = uploaded.type
                    if media_type == "image/jpg": media_type = "image/jpeg"
                    raw = call_claude(api_key, file_bytes, media_type)
                    data = parse_json(raw)
                    st.session_state["result"] = data
                    st.session_state["raw"] = raw
                    st.success("Factura analizada correctamente")
                except Exception as e:
                    st.error("Error: {}".format(e))

        if "result" in st.session_state:
            data = st.session_state["result"]

            c1, c2, c3 = st.columns(3)
            with c1:
                st.download_button("Excel", build_excel_single(data), "factura.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with c2:
                st.download_button("CSV", build_csv_single(data), "factura.csv", "text/csv", use_container_width=True)
            with c3:
                st.download_button("JSON", st.session_state["raw"], "factura.json", "application/json", use_container_width=True)

            emisor  = data.get("emisor", {})
            receptor= data.get("receptor", {})
            factura = data.get("factura", {})
            fields = [
                ("Emisor", emisor.get("nombre","")), ("RUT Emisor", emisor.get("rut","")),
                ("Receptor", receptor.get("nombre","")), ("RUT Receptor", receptor.get("rut","")),
                ("N Factura", factura.get("numero","")), ("Fecha", factura.get("fecha","")),
                ("Cond. Venta", factura.get("condicion_venta","")), ("Vendedor", factura.get("vendedor","")),
            ]
            fields = [(k,v) for k,v in fields if v]
            if fields:
                st.markdown("**Informacion General**")
                cols = st.columns(2)
                for i, (k, v) in enumerate(fields):
                    with cols[i % 2]:
                        st.markdown('<div class="info-card"><div class="info-label">{}</div><div class="info-value">{}</div></div>'.format(k,v), unsafe_allow_html=True)

            items = data.get("items", [])
            if items:
                st.markdown("**Lineas de detalle - {} items**".format(len(items)))
                df_items = pd.DataFrame(items)
                rename = {"codigo":"Codigo","descripcion":"Descripcion","unidad":"U/M",
                          "cantidad":"Cantidad","precio_unitario":"P. Unitario","descuento":"Dto.","total":"Total"}
                df_items = df_items.rename(columns=rename)
                for col in ["P. Unitario","Total","Dto."]:
                    if col in df_items.columns:
                        df_items[col] = df_items[col].apply(fmt_money)
                st.dataframe(df_items, use_container_width=True, hide_index=True)

            t = data.get("totales", {})
            grand = t.get("total", 0)
            if grand:
                st.markdown("**Totales**")
                for label, key in [("Monto Neto","neto"),("Monto Exento","exento"),("IVA 19%","iva")]:
                    val = t.get(key, 0)
                    if val:
                        st.markdown("<div style='display:flex;justify-content:space-between;padding:4px 0'><span>{}</span><span style='font-family:monospace;font-weight:600'>{}</span></div>".format(label, fmt_money(val)), unsafe_allow_html=True)
                st.markdown("<div class='total-grand'><span>TOTAL</span><span>{}</span></div>".format(fmt_money(grand)), unsafe_allow_html=True)

            if data.get("notas"):
                st.warning("{}".format(data["notas"]))

            with st.expander("Ver JSON completo"):
                st.code(st.session_state.get("raw",""), language="json")
        else:
            st.info("Sube una factura y presiona Analizar para comenzar.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: Lote de facturas
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("#### Procesamiento por Lote")
    st.caption("Sube multiples facturas. Se procesaran una a una y se generara una base de datos unificada con una fila por articulo.")

    uploaded_lote = st.file_uploader(
        "Selecciona una o mas facturas",
        type=["jpg","jpeg","png","webp","pdf"],
        accept_multiple_files=True,
        key="lote"
    )

    if uploaded_lote:
        st.markdown("**{} archivo(s) cargado(s)**".format(len(uploaded_lote)))
        for f in uploaded_lote:
            st.markdown("- {}".format(f.name))

    procesar = st.button("Procesar Lote", disabled=not uploaded_lote or not api_key, key="btn_lote")

    if procesar and uploaded_lote and api_key:
        all_rows = []
        errores = []
        progress = st.progress(0, text="Iniciando...")
        status_container = st.container()

        for idx, f in enumerate(uploaded_lote):
            pct = int((idx / len(uploaded_lote)) * 100)
            progress.progress(pct, text="Procesando {} ({}/{})...".format(f.name, idx+1, len(uploaded_lote)))

            try:
                file_bytes = f.read()
                media_type = f.type
                if media_type == "image/jpg": media_type = "image/jpeg"
                raw = call_claude(api_key, file_bytes, media_type)
                data = parse_json(raw)
                rows = factura_to_rows(data, filename=f.name)
                all_rows.extend(rows)
                with status_container:
                    st.markdown('<div class="lote-card lote-ok">✅ <b>{}</b> — {} item(s) extraidos</div>'.format(f.name, len(rows)), unsafe_allow_html=True)
            except Exception as e:
                errores.append((f.name, str(e)))
                with status_container:
                    st.markdown('<div class="lote-card lote-err">❌ <b>{}</b> — Error: {}</div>'.format(f.name, str(e)), unsafe_allow_html=True)

            # Pausa breve para no saturar la API
            if idx < len(uploaded_lote) - 1:
                time.sleep(1)

        progress.progress(100, text="Completado.")

        if all_rows:
            df_lote = pd.DataFrame(all_rows)
            st.session_state["lote_df"] = df_lote
            st.success("{} facturas procesadas, {} filas en total, {} error(es).".format(
                len(uploaded_lote) - len(errores), len(all_rows), len(errores)
            ))

    if "lote_df" in st.session_state:
        df_lote = st.session_state["lote_df"]

        st.markdown("#### Base de Datos Generada")
        st.caption("{} filas x {} columnas".format(len(df_lote), len(df_lote.columns)))
        st.dataframe(df_lote, use_container_width=True, hide_index=True)

        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                "Descargar Excel",
                build_excel_lote(df_lote),
                "facturas_lote.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_lote_xlsx"
            )
        with c2:
            st.download_button(
                "Descargar CSV",
                df_lote.to_csv(index=False).encode("utf-8"),
                "facturas_lote.csv",
                "text/csv",
                use_container_width=True,
                key="dl_lote_csv"
            )
